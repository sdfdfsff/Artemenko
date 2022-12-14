import csv
import datetime
import string
from functools import reduce

from openpyxl.styles import Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.workbook import Workbook

from matplotlib import pyplot as plt

from jinja2 import Environment, FileSystemLoader
import pdfkit
import keyboard
from prettytable import PrettyTable


# <---------------------------------------------------Отчёт-------------------------------------------------------->


clasaas UserInput:
    def __init__(self):
        self.file_name, self.profession = UserInput.get_params()

    @staticmethod
    def get_params():
        file_name = input("Введите название файла: ")
        profession = input("Введите название профессии: ")
        return file_name, profession

    @staticmethod
    def print_data(data, prof):
        salary_by_cities, salary_by_years, vacancies_counts_by_years, vacancies_salary_by_years, vacancies_by_cities, \
        vacancies_by_years = UserInput.get_data(data, prof)

        print("Динамика уровня зарплат по годам:", salary_by_years)
        print("Динамика количества вакансий по годам:", vacancies_by_years)
        print("Динамика уровня зарплат по годам для выбранной профессии:", vacancies_salary_by_years)
        print("Динамика количества вакансий по годам для выбранной профессии:", vacancies_counts_by_years)
        print("Уровень зарплат по городам (в порядке убывания):", salary_by_cities)
        print("Доля вакансий по городам (в порядке убывания):", vacancies_by_cities)

    @staticmethod
    def get_data(data, profession):
        years = set()
        area_dict = {}
        vacancies_dict = {}
        for vac in data:
            years.add(int(datetime.datetime.strptime(vac.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y")))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))
        salary_by_years, vacancies_counts_by_years, vacancies_salary_by_years, vacancies_by_years = UserInput.prepare_dicts(
            years)
        for vacancy in data:
            year = int(datetime.datetime.strptime(vacancy.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
            vacancies_by_years[year] += 1
            salary_by_years[year].append(vacancy.salary.get_salary_ru())
            if profession in vacancy.name:
                vacancies_counts_by_years[year] += 1
                vacancies_salary_by_years[year].append(vacancy.salary.get_salary_ru())
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.get_salary_ru())
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.get_salary_ru()]
            if vacancy.area_name in vacancies_dict:
                vacancies_dict[vacancy.area_name] += 1
            else:
                vacancies_dict[vacancy.area_name] = 1
        salary_by_years, vacancies_salary_by_years = UserInput.set_salaries_by_years(salary_by_years,
                                                                                     vacancies_salary_by_years)
        salary_by_cities = UserInput.set_salaries_by_cities(area_dict, data)
        vacs_by_cities = UserInput.set_vacancies_by_cities(data, vacancies_dict)
        return salary_by_cities, salary_by_years, vacancies_counts_by_years, vacancies_salary_by_years, vacs_by_cities, vacancies_by_years

    @staticmethod
    def prepare_dicts(years):
        salary_by_years = {year: [] for year in years}
        vacs_by_years = {year: 0 for year in years}
        vac_salary_by_years = {year: [] for year in years}
        vac_counts_by_years = {year: 0 for year in years}
        return salary_by_years, vac_counts_by_years, vac_salary_by_years, vacs_by_years

    @staticmethod
    def set_vacancies_by_cities(data, vacs_dict):
        vacs_count = {x: round(y / len(data), 4) for x, y in vacs_dict.items()}
        vacs_count = {k: value for k, value in vacs_count.items() if value >= 0.01}
        vacs_by_cities = dict(sorted(vacs_count.items(), key=lambda value: value[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])
        return vacs_by_cities

    @staticmethod
    def set_salaries_by_cities(area_dict, data):
        area_list = [x for x in area_dict.items() if len(x[1]) / len(data) > 0.01]
        area_list = sorted(area_list, key=lambda value: sum(value[1]) / len(value[1]), reverse=True)
        salary_by_cities = {value[0]: int(sum(value[1]) / len(value[1])) for value in
                            area_list[0: min(len(area_list), 10)]}
        return salary_by_cities

    @staticmethod
    def set_salaries_by_years(salary_by_years, vac_salary_by_years):
        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        vac_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               vac_salary_by_years.items()}
        return salary_by_years, vac_salary_by_years


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.parse_row(file_name)

    @staticmethod
    def сsv_reader(file_name):
        file_csv = open(file_name, encoding="utf_8_sig")
        reader_csv = csv.reader(file_csv)
        list_data = [x for x in reader_csv]
        DataSet.check_file(list_data)
        columns = list_data[0]
        result = [x for x in list_data[1:] if len(x) == len(columns) and x.count('') == 0]
        return columns, result

    @staticmethod
    def check_file(list_data):
        if len(list_data) == 0:
            print("Пустой файл")
            exit()
        if len(list_data) == 1:
            print("Нет данных")
            exit()

    @staticmethod
    def parse_row(file_name):
        name, rows = DataSet.сsv_reader(file_name)
        result = []
        for row in rows:
            new_row = dict(zip(name, row))
            result.append(Vacancy(new_row))
        return result


class Vacancy:
    def __init__(self, row):
        self.name = row["name"]
        self.salary = Salary(row["salary_from"], row["salary_to"], row["salary_currency"])
        self.area_name = row["area_name"]
        self.published_at = row["published_at"]


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_in_rub = int((float(self.salary_from) + float(self.salary_to)) / 2) * currency_to_rub[
            self.salary_currency]

    def get_salary_ru(self):
        return self.salary_in_rub


ru_words = {"name": "Название",
            "description": "Описание",
            "key_skills": "Навыки",
            "experience_id": "Опыт работы",
            "premium": "Премиум-вакансия",
            "employer_name": "Компания",
            "salary": "Оклад",
            "area_name": "Название региона",
            "published_at": "Дата публикации вакансии"}
currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055}


class Report:

    @staticmethod
    def generate_excel(prof):
        workbook = Workbook()
        row_names = list(string.ascii_uppercase)
        headers_font = Font(bold=True)
        border = Border(left=Side(style=BORDER_THIN,
                                  color='00000000'),
                        right=Side(style=BORDER_THIN,
                                   color='00000000'),
                        top=Side(style=BORDER_THIN,
                                 color='00000000'),
                        bottom=Side(style=BORDER_THIN,
                                    color='00000000'))
        max_column_width = {}
        dynamicsForYears = workbook.active
        dynamicsForYears.title = 'Статистика по годам'
        headersForYears = ["Год", "Средняя зарплата", f"Средняя зарпалата - {prof}", "Количество вакансий",
                           f"Количество вакансий - {prof}"]
        for i, header in enumerate(headersForYears):
            dynamicsForYears.cell(row=1, column=i + 1).value = header
            dynamicsForYears[f"{row_names[i]}1"].font = headers_font
        for i, year in enumerate(salary_by_years.keys()):
            dynamicsForYears.cell(row=i + 2, column=1).value = year
            dynamicsForYears.cell(row=i + 2, column=2).value = salary_by_years[year]
            dynamicsForYears.cell(row=i + 2, column=3).value = vacancies_salary_by_years[year]
            dynamicsForYears.cell(row=i + 2, column=4).value = vacancies_by_years[year]
            dynamicsForYears.cell(row=i + 2, column=5).value = vacancies_counts_by_years[year]
        for i, column in enumerate(dynamicsForYears.columns):
            max_column_width[i] = 0
            for cell in column:
                cell.border = border
                if len(str(cell.value)) + 2 > max_column_width[i]:
                    max_column_width[i] = len(str(cell.value)) + 2
            dynamicsForYears.column_dimensions[row_names[i]].width = max_column_width[i]
        dynamicsForCities = workbook.create_sheet(title="Статистика по городам")
        headersForCities = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        for i, header in enumerate(headersForCities):
            dynamicsForCities.cell(row=1, column=i + 1).value = header
            dynamicsForCities[f"{row_names[i]}1"].font = headers_font
        for i, city in enumerate(salary_by_cities.keys()):
            dynamicsForCities.cell(row=i + 2, column=1).value = city
            dynamicsForCities.cell(row=i + 2, column=2).value = salary_by_cities[city]
        for i, city in enumerate(vacs_by_cities.keys()):
            dynamicsForCities.cell(row=i + 2, column=4).value = city
            dynamicsForCities.cell(row=i + 2, column=5).value = f"{round(vacs_by_cities[city] * 100)}%"
        for i, column in enumerate(dynamicsForCities.columns):
            max_column_width[i] = 0
            if i == 2:
                dynamicsForCities.column_dimensions[row_names[3]].width = 2
                continue
            for cell in column:
                cell.border = border
                if len(str(cell.value)) + 2 > max_column_width[i]:
                    max_column_width[i] = len(str(cell.value)) + 2
            dynamicsForCities.column_dimensions[row_names[i]].width = max_column_width[i]
        workbook.save("report.xlsx")

    @staticmethod
    def generate_graphs(prof):
        fig, axs = plt.subplots(2, 2, figsize=(20, 9))
        Report.get_salary_by_years_graph(axs, prof)
        Report.get_salary_by_cities_graph(axs)
        Report.get_vacancies_by_year_graph(axs, prof)
        Report.get_vacancies_parts_by_cities(axs)
        plt.savefig("graph.png")

    @staticmethod
    def get_salary_by_years_graph(axs, prof):
        x_labels = list(salary_by_years.keys())
        x_by_year = list(map(lambda x: x - 0.2, list(salary_by_years.keys())))
        y_by_year = list(salary_by_years.values())

        x_by_profession = list(map(lambda x: x + 0.2, list(vacancies_salary_by_years.keys())))
        y_by_profession = list(vacancies_salary_by_years.values())

        axs[0, 0].title.set_text("Уровень зарплат по годам")
        axs[0, 0].legend(fontsize=8, loc='upper left')
        axs[0, 0].grid(axis='y')
        axs[0, 0].set_xticks(ticks=x_labels, labels=x_labels, rotation=90, fontsize=8)
        plt.tick_params(labelsize=8)
        axs[0, 0].bar(x_by_year, y_by_year, label="средняя з/п", width=0.4)
        axs[0, 0].bar(x_by_profession, y_by_profession,
                      label=f"з/п {prof}", width=0.4)

    @staticmethod
    def get_salary_by_cities_graph(axs):
        x = list(map(lambda x: x.replace("-", "-\n").replace(" ", "\n"), list(salary_by_cities.keys())))
        y = list(salary_by_cities.values())
        x.reverse()
        y.reverse()

        axs[1, 0].title.set_text("Уровень зарплат по городам")
        axs[1, 0].grid(axis="x")
        axs[1, 0].barh(x, y)

    @staticmethod
    def get_vacancies_by_year_graph(axs, prof):
        x_labels = list(vacancies_by_years.keys())

        x_by_year = list(map(lambda x: x - 0.2, list(vacancies_by_years.keys())))
        y_by_year = list(vacancies_by_years.values())

        x_by_prof = list(map(lambda x: x + 0.2, list(vacancies_counts_by_years.keys())))
        y_by_prof = list(vacancies_counts_by_years.values())

        axs[0, 1].title.set_text("Количество вакансий по годам")
        axs[0, 1].legend(fontsize=8, loc='upper left')
        axs[0, 1].grid(axis='y')
        axs[0, 1].bar(x_by_year, y_by_year, label="Количество вакансий", width=0.4)
        axs[0, 1].bar(x_by_prof, y_by_prof,
                      label="Количество вакансий {}".format(prof), width=0.4)
        axs[0, 1].set_xticks(ticks=x_labels, labels=x_labels, rotation=90, fontsize=8)

    @staticmethod
    def get_vacancies_parts_by_cities(axs):
        x = list(vacs_by_cities.keys())
        y = list(vacs_by_cities.values())
        x.insert(0, "Другое")
        y.insert(0, 1 - reduce(lambda X, Y: X + Y, y))

        print(x, y)

        axs[1, 1].title.set_text("Доля вакансий по городам")
        axs[1, 1].pie(y, labels=x, textprops={'fontsize': 6})

    @staticmethod
    def generate_pdf(prof):
        headers1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {prof}", "Количество вакансий",
                    f"Количество вакансий - {prof}"]
        headers2 = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render({'profession': prof,
                                        "headers1": headers1,
                                        "headers2": headers2,
                                        "salary_by_years": salary_by_years,
                                        "vacancies_by_years": vacancies_by_years,
                                        "vacancies_salary_by_years": vacancies_salary_by_years,
                                        "vacancies_counts_by_years": vacancies_counts_by_years,
                                        "salary_by_cities": salary_by_cities,
                                        "vacs_by_cities": vacs_by_cities
                                        })

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


# <---------------------------------------------------Таблицы-------------------------------------------------------->

def get_user_input():
    error_index = -1
    global is_need_to_print
    file_name = input("Введите название файла: ")
    filter_param = input("Введите параметр фильтрации: ")
    sort_param = input("Введите параметр сортировки: ")
    is_reversed_sort = input("Обратный порядок сортировки (Да / Нет): ")
    filter_param, error_index, is_reversed_sort = is_correct_inputs(filter_param, error_index, is_reversed_sort,
                                                                    sort_param)
    table_range = input("Введите диапазон вывода: ").split()
    cols_filter = input("Введите требуемые столбцы: ").split(", ")
    if error_index >= 0:
        print(errors[error_index])
    return file_name, filter_param, sort_param, is_reversed_sort, table_range, cols_filter, is_need_to_print


def is_correct_inputs(filter_param, error_index, is_reverse_sort, sort_parameter):
    global is_need_to_print

    if filter_param != "" and ":" not in filter_param:
        error_index = 0
        is_need_to_print = False
    elif filter_param != "" and not filter_param.split(": ")[0] in filter_types:
        error_index = 1
        is_need_to_print = False
    else:
        filter_param = filter_param.split(": ")

    if sort_parameter not in filter_types:
        error_index = 2
        is_need_to_print = False
    if is_reverse_sort not in reverse_parameter:
        error_index = 3
        is_need_to_print = False
    else:
        is_reverse_sort = reverse_parameter[is_reverse_sort]

    return filter_param, error_index, is_reverse_sort


def check_skills(row, words):
    current_count = 0
    words = words.split(", ")
    check_count = len(words)
    row = row.split("*- ")
    for x in words:
        if x in row:
            current_count += 1
    return current_count == check_count


def filter_row(row, words_filter):
    key, value = "", ""
    if words_filter != "":
        for k, v in words_filter.items():
            key = k
            value = v
    is_need_to_print = False
    if key in list(filter_types.keys()):
        is_need_to_print = filter_types[key](row, value)
    if is_need_to_print:
        return row
    return


def check_inputs(table):
    if len(table_range) >= 1 and int(table_range[0]) <= len(table.rows):
        start_row = int(table_range[0]) - 1
    else:
        start_row = 0
    if len(table_range) == 2 and int(table_range[0]) <= len(table.rows) and int(table_range[1]) >= int(table_range[0]):
        end_row = int(table_range[1]) - 1
    else:
        end_row = len(table.rows)
    if cols_filter[0] != '':
        titles = ["№"] + cols_filter
    else:
        titles = table.field_names
    return start_row, end_row, titles


def format_row(row):
    formatted_row = {}
    for x in ru_words:
        if x in functions:
            formatted_row[ru_words[x]] = functions[x](row)
        else:
            formatted_row[ru_words[x]] = row[x]
        if len(formatted_row[ru_words[x]]) >= 102:
            formatted_row[ru_words[x]] = functions["check_length"](formatted_row, ru_words[x])
    return formatted_row


def parse_row(row, name):
    parsed_row = dict(zip(name, row))
    for x in parsed_row:
        if x == "name":
            parsed_row[x] = " ".join(parsed_row[x].split())
        if x == "description":
            parsed_row[x] = " ".join((re.sub(r'\<[^>]*\>', '', parsed_row[x])).split())
        elif x == "key_skills":
            parsed_row[x] = parsed_row[x].replace("\n", "*- ")
        elif x == "premium" or x == "salary_gross":
            parsed_row[x] = parsed_row[x].replace("True", "Да")
            parsed_row[x] = parsed_row[x].replace("False", "Нет")
    return parsed_row


def universal_parser(f):
    def wrapper(name):
        res = f(name)
        if res is not None:
            print_vacancies(res, ru_words)

    return wrapper


@universal_parser
def сsv_reader(file_name):
    file_csv = open(file_name, encoding="utf_8_sig")
    list_data = [x for x in csv.reader(file_csv)]
    if len(list_data) == 0:
        print("Пустой файл")
        columns, result = [], ["error"]
    elif len(list_data) == 1:
        print("Нет данных")
        columns, result = list_data[0], ["error"]
    else:
        columns = list_data[0]
        result = [x for x in list_data[1:] if len(x) == len(columns) and x.count('') == 0]
    if result[0] != "error" and is_need_to_print:
        is_empty = len(result) == 0
        for x in range(len(result)):
            result[x] = filter_row(parse_row(result[x], columns), filter_param)
        result = [x for x in result if x is not None]
        if len(result) == 0 and is_empty is False:
            print("Ничего не найдено")
        else:
            result = sort_types[sort_parameter](result, is_reverse_sort)
        return result


def print_vacancies(data_vacancies, ru_words):
    if len(data_vacancies) == 0:
        return
    table = PrettyTable()
    table._max_width = {x: 20 for x in (["№"] + list(ru_words.values()))}
    table.field_names = ["№"] + list(ru_words.values())
    table.align = "l"
    table.hrules = True
    i = 0
    for x in range(len(data_vacancies)):
        i += 1
        formatted = list(format_row(data_vacancies[x]).values())
        formatted.insert(0, x + 1)
        table.add_row(formatted)
    start_row, end_row, titles, = check_inputs(table)
    if is_need_to_print:
        print(table.get_string(start=start_row, end=end_row, fields=titles))

def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k

errors = ["Формат ввода некорректен",
          "Параметр поиска некорректен",
          "Параметр сортировки некорректен",
          "Порядок сортировки задан некорректно"]
salary_gross_dic = {"Нет": "С вычетом налогов",
                    "Да": "Без вычета налогов"}

ru_words = {"name": "Название",
            "description": "Описание",
            "key_skills": "Навыки",
            "experience_id": "Опыт работы",
            "premium": "Премиум-вакансия",
            "employer_name": "Компания",
            "a": "Оклад",
            "area_name": "Название региона",
            "published_at": "Дата публикации вакансии"}
ru_exp = {"noExperience": "Нет опыта",
          "between1And3": "От 1 года до 3 лет",
          "between3And6": "От 3 до 6 лет",
          "moreThan6": "Более 6 лет"}
ru_currency = {"AZN": "Манаты",
               "BYR": "Белорусские рубли",
               "EUR": "Евро",
               "GEL": "Грузинский лари",
               "KGS": "Киргизский сом",
               "KZT": "Тенге",
               "RUR": "Рубли",
               "UAH": "Гривны",
               "USD": "Доллары",
               "UZS": "Узбекский сум"}
currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055}
sort_exp = {"noExperience": 0,
            "between1And3": 1,
            "between3And6": 2,
            "moreThan6": 3}
filter_types = {"Название": lambda row, words: row["name"] == words,
                "Описание": lambda row, words: row["description"] == words,
                "Навыки": lambda row, words: check_skills(row["key_skills"], words),
                "Опыт работы": lambda row, words: get_key(ru_exp, words) == row["experience_id"],
                "Премиум-вакансия": lambda row, words: words in row["premium"],
                "Компания": lambda row, words: row["employer_name"] == words,
                "Идентификатор валюты оклада": lambda row, words: get_key(ru_currency, words) == row["salary_currency"],
                "Оклад": lambda row, words: int(float(row["salary_from"])) <= int(words) <= int(
                    float(row["salary_to"])),
                "Название региона": lambda row, words: words in row["area_name"],
                "Дата публикации вакансии": lambda row, words: functions["published_at"](row) == words,
                "": lambda row, words: row}
sort_types = {"Название": lambda row, revers: sorted(row, key=lambda d: d["name"], reverse=revers),
                "Описание": lambda row, revers: sorted(row, key=lambda d: d["description"], reverse=revers),
                "Навыки": lambda row, revers: sorted(row, key=lambda d: len(d["key_skills"].split("*- ")),
                                                     reverse=revers),
                "Опыт работы": lambda row, revers: sorted(row, key=lambda d: sort_exp[d["experience_id"]],
                                                          reverse=revers),
                "Премиум-вакансия": lambda row, revers: sorted(row, key=lambda d: d["premium"], reverse=revers),
                "Компания": lambda row, revers: sorted(row, key=lambda d: d["employer_name"], reverse=revers),
                "Идентификатор валюты оклада": lambda row, revers: sorted(row, key=lambda d:
                ru_currency[d["salary_currency"]], reverse=revers),
                "Оклад": lambda row, revers: sorted(row, key=lambda d:
                (int(float(d["salary_from"])) + int(float(d["salary_to"]))) / 2 * currency_to_rub[d["salary_currency"]],
                                                    reverse=revers),
              "Название региона": lambda row, revers: sorted(row, key=lambda d: d["area_name"], reverse=revers),
              "Дата публикации вакансии": lambda row, revers: sorted(row, key=lambda d: functions["for_sort_date"](d),
                                                                       reverse=revers),
              "": lambda row, words: row}
reverse_parameter = {"Нет": False,
                     "Да": True,
                     "": False}
functions = {"key_skills": lambda row: row["key_skills"].replace("*- ", "\n"),
             "a": lambda row: "{0:,} - {1:,} ({2}) ({3})".format(int(float(row["salary_from"])),
                                                                 int(float(row["salary_to"])),
                                                                 ru_currency[row["salary_currency"]],
                                                                 salary_gross_dic[row["salary_gross"]]).replace(",",
                                                                                                                " "),
             "experience_id": lambda row: ru_exp[row["experience_id"]],
             "published_at": lambda row: datetime.datetime.strptime(row["published_at"], "%Y-%m-%dT%H:%M:%S%z")
             .strftime("%d.%m.%Y"),
             "for_sort_date": lambda row: datetime.datetime.strptime(row["published_at"], "%Y-%m-%dT%H:%M:%S%z")
             .strftime("%Y.%m.%d.%H.%M.%S"),
             "check_length": lambda row, name: row[name][0:100] + " ..." if row[name][101] == " " and row[name][
                 100] == " " else row[name][
                                  0:100] + "..."
             }

print("Выберите действие и нажмите кнопку:")
print("1. Показать данные в виде таблицы.")
print("2. Распечатать отчёт в виде PDF")

while True:
    if keyboard.is_pressed('1'):
        table_range = []
        cols_filter = ['']
        filter_param = ""
        is_need_to_print = True

        file_name, filter_param, sort_parameter, is_reverse_sort, table_range, cols_filter, is_need_to_print = get_user_input()
        if is_need_to_print:
            filter_param = {filter_param[x]: filter_param[x + 1] for x in range(len(filter_param) - 1)}
            сsv_reader(file_name)
        break
    elif keyboard.is_pressed('2'):
        params = UserInput()
        vacs = DataSet(params.file_name).vacancies_objects
        params.print_data(vacs, params.profession)
        salary_by_cities, salary_by_years, vacancies_counts_by_years, vacancies_salary_by_years, vacs_by_cities, \
        vacancies_by_years = UserInput.get_data(vacs, params.profession)
        Report.generate_excel(params.profession)
        Report.generate_graphs(params.profession)
        Report.generate_pdf(params.profession)
        break
